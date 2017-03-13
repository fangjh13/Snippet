#!/usr/bin/evn python3
# -*- coding: utf-8 -*-

'''需要安装openpyxl模块操作xlsx文件，生成result.xlsx文件'''

from openpyxl import load_workbook
from sys import argv


# 根据设备端安装管理文件生成dict
def create_db(sheet):
    n = 3
    db = {}
    while True:
        n += 1
        if fs_db['A{}'.format(n)].value:
            db[str(fs_db['A{}'.format(n)].value)] = \
                {'name': fs_db['B{}'.format(n)].value,
                 'model': fs_db['C{}'.format(n)].value,
                 'version': fs_db['E{}'.format(n)].value,
                 'is_newest': fs_db['F{}'.format(n)].value}
        else:
            break
    return db


# 写入新xlsx函数
def operate(fs, db):
    fs['R5'] = "售货机名称"
    fs['S5'] = "售货机型号"
    fs['T5'] = "当前版本号"
    fs['U5'] = "是否最新版本"

    col = 5
    while True:
        col += 1
        m_id = fs['C{}'.format(col)].value
        if m_id:
            # 统一转换成str避免混乱
            m_id = str(m_id)
            if m_id in db:
                fs['R{}'.format(col)] = db[m_id]['name']
                fs['S{}'.format(col)] = db[m_id]['model']
                fs['T{}'.format(col)] = db[m_id]['version']
                fs['U{}'.format(col)] = db[m_id]['is_newest']
            else:
                print('FATAL: 机器编号：{} 不在设备端安装管理数据库 '.format(m_id))
        else:
            break


if __name__ == '__main__':
    # 设备端安装管理xlsx文件
    db_name = argv[1]
    # 机器收货明细xlsx文件
    detail = argv[2]

    print('数据处理中，请耐心等待。。。。')

    wb = load_workbook(detail)
    wb_db = load_workbook(db_name)

    fs = wb['First Sheet']
    fs_db = wb_db['First Sheet']

    db = create_db(fs_db)

    operate(fs, db)

    print('保存中。。。')

    wb.save('result.xlsx')
    print('操作完成')
